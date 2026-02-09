#!/usr/bin/env python3
"""
MECM Dashboard Excel Workbook Generator

Generates a multi-sheet Excel dashboard workbook with:
- SQL Server ODBC data connections (MECMServer / CM_PS1)
- Pie charts and bar charts per sheet
- SQL queries sourced from the project's .sql files
- Dashboard-specific aggregation queries for Security
- A reference sheet listing all embedded queries
"""

import os
import zipfile
import tempfile
import shutil
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Configuration ────────────────────────────────────────────────────────────
SERVER_NAME = "MECMServer"
DATABASE_NAME = "CM_PS1"
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "MECM_Dashboard.xlsx")
CONN_STRING = (
    f"Provider=SQLOLEDB.1;Data Source={SERVER_NAME};"
    f"Initial Catalog={DATABASE_NAME};Integrated Security=SSPI;"
)

# ── Colors ───────────────────────────────────────────────────────────────────
C = {
    "navy": "1F3864",
    "blue": "2F5496",
    "med_blue": "4472C4",
    "light_blue": "B4C6E7",
    "pale_blue": "D6E4F0",
    "green": "548235",
    "orange": "ED7D31",
    "gold": "FFC000",
    "red": "C00000",
    "gray": "808080",
    "light_gray": "D9D9D9",
    "white": "FFFFFF",
    "dark": "1F1F1F",
}
PIE_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9B57A1", "636363", "FF6B6B",
    "44546A", "BDD7EE", "F4B183", "C5E0B4",
]

# ── Styles ───────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Segoe UI", bold=True, size=18, color=C["navy"])
SUBTITLE_FONT = Font(name="Segoe UI", bold=True, size=12, color=C["blue"])
HEADER_FONT = Font(name="Segoe UI", bold=True, size=10, color=C["white"])
HEADER_FILL = PatternFill(start_color=C["blue"], end_color=C["blue"], fill_type="solid")
DATA_FONT = Font(name="Segoe UI", size=10, color=C["dark"])
ALT_FILL = PatternFill(start_color=C["pale_blue"], end_color=C["pale_blue"], fill_type="solid")
QUERY_FONT = Font(name="Consolas", size=9, color=C["dark"])
THIN_BORDER = Border(
    bottom=Side(style="thin", color=C["navy"]),
    top=Side(style="thin", color=C["navy"]),
)


# ── SQL Queries ──────────────────────────────────────────────────────────────
def read_sql(relative_path):
    """Read a SQL file from the project root."""
    with open(os.path.join(BASE_DIR, relative_path)) as f:
        content = f.read()
    # Strip the comment header, return just the SQL
    lines = content.strip().splitlines()
    sql_lines = []
    in_comment = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("/*"):
            in_comment = True
            continue
        if in_comment:
            if "*/" in stripped:
                in_comment = False
            continue
        sql_lines.append(line)
    return "\n".join(sql_lines).strip()


# Dashboard-specific aggregation queries for Security
SECURITY_QUERIES = {
    "BitLocker Protection Summary": """SELECT
    CASE bd.ProtectionStatus0
        WHEN 0 THEN 'Protection Off'
        WHEN 1 THEN 'Protection On'
        ELSE 'Unknown'
    END AS [Protection Status],
    COUNT(DISTINCT sys.ResourceID) AS [Device Count]
FROM v_R_System_Valid sys
LEFT JOIN v_GS_BITLOCKER_DETAILS bd
    ON sys.ResourceID = bd.ResourceID
    AND bd.DriveLetter0 = 'C:'
GROUP BY bd.ProtectionStatus0
ORDER BY [Device Count] DESC""",
    "Secure Boot Summary": """SELECT
    CASE fw.SecureBoot0
        WHEN 1 THEN 'Enabled'
        WHEN 0 THEN 'Disabled'
        ELSE 'Unknown'
    END AS [Secure Boot Status],
    COUNT(*) AS [Device Count]
FROM v_R_System_Valid sys
LEFT JOIN v_GS_FIRMWARE fw
    ON sys.ResourceID = fw.ResourceID
GROUP BY fw.SecureBoot0
ORDER BY [Device Count] DESC""",
    "Defender Real-Time Protection Summary": """SELECT
    CASE wds.RealTimeProtectionEnabled0
        WHEN 1 THEN 'Enabled'
        WHEN 0 THEN 'Disabled'
        ELSE 'Unknown'
    END AS [Real-Time Protection],
    COUNT(*) AS [Device Count]
FROM v_R_System_Valid sys
LEFT JOIN v_GS_WINDOWS_DEFENDER_STATUS wds
    ON sys.ResourceID = wds.ResourceID
GROUP BY wds.RealTimeProtectionEnabled0
ORDER BY [Device Count] DESC""",
    "TPM Status Summary": """SELECT
    CASE tpm.IsActivated_InitialValue0
        WHEN 1 THEN 'Activated'
        WHEN 0 THEN 'Not Activated'
        ELSE 'Unknown'
    END AS [TPM Status],
    COUNT(*) AS [Device Count]
FROM v_R_System_Valid sys
LEFT JOIN v_GS_TPM tpm
    ON sys.ResourceID = tpm.ResourceID
GROUP BY tpm.IsActivated_InitialValue0
ORDER BY [Device Count] DESC""",
}


# ── Helper Functions ─────────────────────────────────────────────────────────
def add_title(ws, title, subtitle=None, row=1):
    """Add sheet title and optional subtitle. Returns next available row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 35
    row += 1
    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=subtitle)
        cell.font = SUBTITLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1
    return row + 1


def add_section(ws, label, row):
    """Add a section label. Returns next row."""
    ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
    return row + 1


def write_table(ws, headers, data, start_row, start_col=1, name=None):
    """Write a formatted data table. Returns the row after the last data row."""
    for ci, h in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for ri, row_data in enumerate(data, start=start_row + 1):
        for ci, val in enumerate(row_data, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 0:
                cell.fill = ALT_FILL

    # Auto-width columns
    for ci in range(start_col, start_col + len(headers)):
        hi = ci - start_col
        max_len = len(str(headers[hi]))
        for rd in data:
            if hi < len(rd):
                max_len = max(max_len, len(str(rd[hi])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

    end_row = start_row + len(data)

    if name and data:
        end_col = get_column_letter(start_col + len(headers) - 1)
        ref = f"{get_column_letter(start_col)}{start_row}:{end_col}{end_row}"
        tbl = Table(displayName=name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)

    return end_row + 1


def make_pie(ws, title, cats_col, data_col, min_row, max_row, anchor, w=13, h=10):
    """Create a pie chart on the worksheet."""
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = w
    chart.height = h

    data_ref = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
    cats_ref = Reference(ws, min_col=cats_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False
    chart.dataLabels.showSerName = False

    if chart.series:
        s = chart.series[0]
        for i, color in enumerate(PIE_COLORS):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            s.data_points.append(pt)

    ws.add_chart(chart, anchor)


def make_bar(ws, title, cats_col, data_cols, min_row, max_row, anchor,
             w=16, h=10, y_title=None):
    """Create a bar chart. data_cols can be a single int or list of ints."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.width = w
    chart.height = h
    if y_title:
        chart.y_axis.title = y_title

    if isinstance(data_cols, int):
        data_cols = [data_cols]

    for i, dc in enumerate(data_cols):
        ref = Reference(ws, min_col=dc, min_row=min_row, max_row=max_row)
        chart.add_data(ref, titles_from_data=True)

    cats_ref = Reference(ws, min_col=cats_col, min_row=min_row + 1, max_row=max_row)
    chart.set_categories(cats_ref)

    for i, s in enumerate(chart.series):
        s.graphicalProperties.solidFill = PIE_COLORS[i % len(PIE_COLORS)]

    ws.add_chart(chart, anchor)


# ── Sheet Builders ───────────────────────────────────────────────────────────
def build_os_dashboard(wb):
    ws = wb.create_sheet("OS Dashboard")
    ws.sheet_properties.tabColor = C["blue"]

    r = add_title(ws, "Operating System Dashboard",
                  "Distribution of Windows operating systems across managed devices")

    # ── OS Count Summary ──
    r = add_section(ws, "OS Distribution", r)
    headers = ["Operating System", "Device Count"]
    data = [
        ("Microsoft Windows 11 Enterprise", 450),
        ("Microsoft Windows 10 Enterprise", 320),
        ("Microsoft Windows 11 Pro", 150),
        ("Microsoft Windows 10 Pro", 80),
        ("Microsoft Windows Server 2022 Standard", 45),
        ("Microsoft Windows Server 2019 Standard", 30),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="OSCount")
    make_pie(ws, "OS Distribution", 1, 2, t_start, t_start + len(data), f"E{t_start}")

    r += 1

    # ── Feature Update Counts ──
    r = add_section(ws, "Feature Update Distribution", r)
    headers2 = ["Operating System", "Feature Update", "Build Number", "Device Count"]
    data2 = [
        ("Microsoft Windows 11 Enterprise", "24H2", "26100", 280),
        ("Microsoft Windows 11 Enterprise", "23H2", "22631", 170),
        ("Microsoft Windows 10 Enterprise", "22H2", "19045", 320),
        ("Microsoft Windows 11 Pro", "24H2", "26100", 100),
        ("Microsoft Windows 11 Pro", "23H2", "22631", 50),
        ("Microsoft Windows 10 Pro", "22H2", "19045", 80),
    ]
    t2_start = r
    r = write_table(ws, headers2, data2, r, name="FeatureUpdates")
    make_bar(ws, "Feature Update Distribution", 2, 4, t2_start, t2_start + len(data2),
             f"F{t2_start}", y_title="Device Count")


def build_hardware_dashboard(wb):
    ws = wb.create_sheet("Hardware Dashboard")
    ws.sheet_properties.tabColor = C["green"]

    r = add_title(ws, "Hardware Inventory Dashboard",
                  "Memory capacity and device model distribution")

    # ── Memory Summary ──
    r = add_section(ws, "Memory Distribution", r)
    headers = ["RAM (GB)", "Device Count"]
    data = [
        (4, 25),
        (8, 180),
        (16, 450),
        (32, 280),
        (64, 65),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="MemoryDist")
    make_pie(ws, "RAM Distribution", 1, 2, t_start, t_start + len(data), f"E{t_start}")

    r += 1

    # ── Device Models ──
    r = add_section(ws, "Top Device Models", r)
    headers2 = ["Manufacturer", "Model", "Chassis Type", "Device Count"]
    data2 = [
        ("Dell Inc.", "Latitude 5540", "Laptop", 145),
        ("Dell Inc.", "OptiPlex 7090", "Desktop", 130),
        ("Lenovo", "ThinkPad T14 Gen 4", "Notebook", 115),
        ("HP", "EliteBook 840 G10", "Notebook", 95),
        ("Dell Inc.", "Latitude 7440", "Laptop", 85),
        ("Lenovo", "ThinkCentre M90q", "Desktop", 70),
        ("Microsoft Corporation", "Virtual Machine", "Other", 60),
        ("HP", "ProDesk 400 G7", "Small Form Factor", 55),
    ]
    t2_start = r
    r = write_table(ws, headers2, data2, r, name="DeviceModels")
    make_bar(ws, "Top Device Models", 2, 4, t2_start, t2_start + len(data2),
             f"F{t2_start}", y_title="Count")


def build_client_health(wb):
    ws = wb.create_sheet("Client Health")
    ws.sheet_properties.tabColor = C["orange"]

    r = add_title(ws, "Client Health Dashboard",
                  "MECM client version distribution and health status")

    # ── Client Version ──
    r = add_section(ws, "Client Version Distribution", r)
    headers = ["Client Version", "Device Count"]
    data = [
        ("5.00.9128.1007", 520),
        ("5.00.9122.1009", 280),
        ("5.00.9114.1012", 120),
        ("5.00.9106.1000", 45),
        ("No Client", 35),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="ClientVersion")
    make_pie(ws, "MECM Client Versions", 1, 2, t_start, t_start + len(data), f"E{t_start}")


def build_update_compliance(wb):
    ws = wb.create_sheet("Update Compliance")
    ws.sheet_properties.tabColor = C["gold"]

    r = add_title(ws, "Software Update Compliance Dashboard",
                  "Patch compliance and deployment status across update classifications")

    # ── Compliance Summary ──
    r = add_section(ws, "Compliance by Classification", r)
    headers = ["Classification", "Total Devices", "Required", "Installed", "Compliance %"]
    data = [
        ("Security Updates", 1000, 150, 850, 85.0),
        ("Critical Updates", 1000, 50, 950, 95.0),
        ("Definition Updates", 1000, 200, 800, 80.0),
        ("Update Rollups", 1000, 100, 900, 90.0),
        ("Feature Packs", 800, 40, 760, 95.0),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="ComplianceSummary")
    make_bar(ws, "Compliance % by Classification", 1, 5, t_start, t_start + len(data),
             f"G{t_start}", y_title="Compliance %")

    r += 1

    # ── Deployment Status ──
    r = add_section(ws, "Update Group Deployment Status", r)
    headers2 = ["Update Group", "Targeted Devices", "Compliant", "Required",
                "Not Required", "Compliance %"]
    data2 = [
        ("2026-01 Security Updates", 1000, 920, 50, 30, 92.0),
        ("2026-02 Security Updates", 1000, 750, 200, 50, 75.0),
        (".NET Framework Updates", 800, 780, 10, 10, 97.5),
        ("Microsoft 365 Updates", 900, 810, 60, 30, 90.0),
    ]
    t2_start = r
    r = write_table(ws, headers2, data2, r, name="DeploymentStatus")
    make_bar(ws, "Deployment Compliance by Update Group", 1, 6, t2_start,
             t2_start + len(data2), f"H{t2_start}", y_title="Compliance %")


def build_security_dashboard(wb):
    ws = wb.create_sheet("Security Dashboard")
    ws.sheet_properties.tabColor = C["red"]

    r = add_title(ws, "Security Dashboard",
                  "BitLocker, Secure Boot, TPM, and Defender protection status")

    # ── BitLocker ──
    r = add_section(ws, "BitLocker Protection Status", r)
    headers = ["Protection Status", "Device Count"]
    data = [
        ("Protection On", 820),
        ("Protection Off", 130),
        ("Unknown", 50),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="BitLockerStatus")
    make_pie(ws, "BitLocker Protection", 1, 2, t_start, t_start + len(data), f"E{t_start}")

    r += 1

    # ── Secure Boot ──
    r = add_section(ws, "Secure Boot Status", r)
    headers2 = ["Secure Boot Status", "Device Count"]
    data2 = [
        ("Enabled", 850),
        ("Disabled", 100),
        ("Unknown", 50),
    ]
    t2_start = r
    r = write_table(ws, headers2, data2, r, name="SecureBootStatus")
    make_pie(ws, "Secure Boot", 1, 2, t2_start, t2_start + len(data2), f"E{t2_start}")

    r += 1

    # ── Defender Real-Time Protection ──
    r = add_section(ws, "Defender Real-Time Protection", r)
    headers3 = ["Real-Time Protection", "Device Count"]
    data3 = [
        ("Enabled", 940),
        ("Disabled", 35),
        ("Unknown", 25),
    ]
    t3_start = r
    r = write_table(ws, headers3, data3, r, name="DefenderRTP")
    make_pie(ws, "Defender Real-Time Protection", 1, 2, t3_start, t3_start + len(data3),
             f"E{t3_start}")

    r += 1

    # ── TPM ──
    r = add_section(ws, "TPM Activation Status", r)
    headers4 = ["TPM Status", "Device Count"]
    data4 = [
        ("Activated", 900),
        ("Not Activated", 60),
        ("Unknown", 40),
    ]
    t4_start = r
    r = write_table(ws, headers4, data4, r, name="TPMStatus")
    make_pie(ws, "TPM Activation", 1, 2, t4_start, t4_start + len(data4), f"E{t4_start}")


def build_applications_dashboard(wb):
    ws = wb.create_sheet("Applications")
    ws.sheet_properties.tabColor = "9B57A1"

    r = add_title(ws, "Application Deployment Dashboard",
                  "Deployment summary and per-application success rates")

    # ── Deployment Summary (single row → pie chart of success/fail/progress) ──
    r = add_section(ws, "Overall Deployment Summary", r)
    headers = ["Metric", "Count"]
    data = [
        ("Successful", 8500),
        ("Failed", 320),
        ("In Progress", 180),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="DeploySummary")
    make_pie(ws, "Deployment Outcome Distribution", 1, 2,
             t_start, t_start + len(data), f"E{t_start}")

    r += 1

    # ── Per-App Deployment Status ──
    r = add_section(ws, "Application Deployment Status", r)
    headers2 = ["Application Name", "Manufacturer", "Target Collection",
                "Total Targeted", "Success", "In Progress", "Errors", "Success Rate %"]
    data2 = [
        ("Microsoft 365 Apps", "Microsoft", "All Workstations", 1000, 950, 20, 30, 95.0),
        ("Google Chrome", "Google", "All Workstations", 1000, 980, 10, 10, 98.0),
        ("Adobe Acrobat Reader", "Adobe", "All Workstations", 800, 750, 25, 25, 93.8),
        ("Zoom Workplace", "Zoom", "All Workstations", 900, 860, 15, 25, 95.6),
        ("7-Zip", "Igor Pavlov", "All Workstations", 700, 695, 0, 5, 99.3),
        ("Notepad++", "Don Ho", "Developer PCs", 200, 195, 3, 2, 97.5),
    ]
    t2_start = r
    r = write_table(ws, headers2, data2, r, name="AppDeployStatus")
    make_bar(ws, "Success Rate by Application", 1, 8, t2_start, t2_start + len(data2),
             f"J{t2_start}", y_title="Success Rate %")


def build_server_dashboard(wb):
    ws = wb.create_sheet("Server Dashboard")
    ws.sheet_properties.tabColor = "44546A"

    r = add_title(ws, "Server Dashboard",
                  "Windows Server OS version distribution")

    # ── Server OS Versions ──
    r = add_section(ws, "Server OS Distribution", r)
    headers = ["Operating System", "Server Version", "Build Number", "Server Count"]
    data = [
        ("Microsoft Windows Server 2022 Standard", "Server 2022", "20348", 35),
        ("Microsoft Windows Server 2019 Standard", "Server 2019", "17763", 28),
        ("Microsoft Windows Server 2022 Datacenter", "Server 2022", "20348", 12),
        ("Microsoft Windows Server 2016 Standard", "Server 2016", "14393", 10),
        ("Microsoft Windows Server 2019 Datacenter", "Server 2019", "17763", 8),
        ("Microsoft Windows Server 2012 R2 Standard", "Server 2012 R2", "9600", 5),
    ]
    t_start = r
    r = write_table(ws, headers, data, r, name="ServerOS")
    make_pie(ws, "Server OS Distribution", 2, 4, t_start, t_start + len(data), f"F{t_start}")

    # Also a bar chart by server version
    make_bar(ws, "Server Count by Version", 2, 4, t_start, t_start + len(data),
             f"F{t_start + 16}", y_title="Server Count")


def build_queries_sheet(wb, query_map):
    """Reference sheet listing all SQL queries used in the workbook."""
    ws = wb.create_sheet("SQL Queries")
    ws.sheet_properties.tabColor = C["gray"]

    r = add_title(ws, "SQL Queries Reference",
                  "All queries used by this dashboard workbook")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 100

    r = add_section(ws, "Query Index", r)

    # Header row
    for ci, h in enumerate(["Query Name", "Source File", "SQL Query"], start=1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    for name, info in query_map.items():
        ws.cell(row=r, column=1, value=name).font = DATA_FONT
        ws.cell(row=r, column=2, value=info.get("source", "Dashboard")).font = DATA_FONT
        cell = ws.cell(row=r, column=3, value=info["sql"])
        cell.font = QUERY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(60, len(info["sql"]) // 2)
        r += 1


def build_connection_sheet(wb):
    """Setup instructions for the data source connection."""
    ws = wb.create_sheet("Connection Setup")
    ws.sheet_properties.tabColor = C["navy"]

    r = add_title(ws, "Data Source Connection Setup",
                  "Configure the SQL Server connection to your MECM database")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    params = [
        ("Server Name", SERVER_NAME),
        ("Database Name", DATABASE_NAME),
        ("Authentication", "Windows Authentication (Integrated Security)"),
        ("Provider", "SQLOLEDB.1"),
        ("Connection String", CONN_STRING),
    ]

    r = add_section(ws, "Connection Parameters", r)
    for label, value in params:
        ws.cell(row=r, column=1, value=label).font = Font(
            name="Segoe UI", bold=True, size=10, color=C["dark"])
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = Font(name="Consolas", size=10, color=C["blue"])
        cell.alignment = Alignment(wrap_text=True)
        r += 1

    r += 1
    r = add_section(ws, "How to Refresh Data", r)

    steps = [
        "1. Open this workbook in Excel (desktop version).",
        "2. Go to Data > Connections to see the embedded data connections.",
        f"3. Update the server name from '{SERVER_NAME}' to your actual MECM SQL server.",
        f"4. Update the database name from '{DATABASE_NAME}' to your site code (e.g., CM_ABC).",
        "5. Click Data > Refresh All to execute all queries against your database.",
        "6. Charts will automatically update with live data from your MECM environment.",
        "",
        "Note: You must have read access to the MECM site database.",
        "The connection uses Windows Authentication (your logged-in credentials).",
        "If using a named instance, set server to: ServerName\\InstanceName",
        "",
        "Alternative: Use Power Query (Get & Transform Data) for more control.",
        "Copy the SQL from the 'SQL Queries' sheet into Power Query's native query editor.",
    ]
    for step in steps:
        ws.cell(row=r, column=1, value=step).font = DATA_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1


# ── OOXML Connection Injection ───────────────────────────────────────────────
def inject_connections(xlsx_path, connections):
    """
    Inject ODBC connection definitions into the .xlsx zip structure.
    connections: list of dicts with 'id', 'name', 'sql' keys.
    """
    tmp = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(xlsx_path, "r") as z:
            z.extractall(tmp)

        ns_ss = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns_ct = "http://schemas.openxmlformats.org/package/2006/content-types"
        ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"

        ET.register_namespace("", ns_ss)

        # 1. Create xl/connections.xml
        root = ET.Element(f"{{{ns_ss}}}connections")
        for c in connections:
            elem = ET.SubElement(root, f"{{{ns_ss}}}connection")
            elem.set("id", str(c["id"]))
            elem.set("name", c["name"])
            elem.set("type", "1")
            elem.set("refreshedVersion", "0")
            elem.set("background", "1")
            elem.set("saveData", "1")

            dbpr = ET.SubElement(elem, f"{{{ns_ss}}}dbPr")
            dbpr.set("connection", CONN_STRING)
            dbpr.set("command", c["sql"])
            dbpr.set("commandType", "2")

        conn_path = os.path.join(tmp, "xl", "connections.xml")
        ET.ElementTree(root).write(conn_path, xml_declaration=True, encoding="UTF-8")

        # 2. Update [Content_Types].xml
        ct_path = os.path.join(tmp, "[Content_Types].xml")
        ET.register_namespace("", ns_ct)
        ct_tree = ET.parse(ct_path)
        ct_root = ct_tree.getroot()

        override = ET.SubElement(ct_root, f"{{{ns_ct}}}Override")
        override.set("PartName", "/xl/connections.xml")
        override.set(
            "ContentType",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
        )
        ct_tree.write(ct_path, xml_declaration=True, encoding="UTF-8")

        # 3. Update xl/_rels/workbook.xml.rels
        rels_path = os.path.join(tmp, "xl", "_rels", "workbook.xml.rels")
        ET.register_namespace("", ns_rel)
        rels_tree = ET.parse(rels_path)
        rels_root = rels_tree.getroot()

        existing_ids = []
        for rel in rels_root:
            rid = rel.get("Id", "")
            if rid.startswith("rId"):
                try:
                    existing_ids.append(int(rid[3:]))
                except ValueError:
                    pass
        next_id = max(existing_ids) + 1 if existing_ids else 1

        rel_elem = ET.SubElement(rels_root, f"{{{ns_rel}}}Relationship")
        rel_elem.set("Id", f"rId{next_id}")
        rel_elem.set(
            "Type",
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/connections",
        )
        rel_elem.set("Target", "connections.xml")
        rels_tree.write(rels_path, xml_declaration=True, encoding="UTF-8")

        # Repackage
        os.remove(xlsx_path)
        with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as z:
            for dirpath, dirnames, filenames in os.walk(tmp):
                for fn in filenames:
                    full = os.path.join(dirpath, fn)
                    arc = os.path.relpath(full, tmp)
                    z.write(full, arc)
    finally:
        shutil.rmtree(tmp)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Read project SQL files
    sql_files = {
        "OS Count Summary": {
            "source": "Operating-Systems/OS_Count_Summary.sql",
            "sql": read_sql("Operating-Systems/OS_Count_Summary.sql"),
        },
        "OS Feature Update Counts": {
            "source": "Operating-Systems/OS_FeatureUpdate_Counts.sql",
            "sql": read_sql("Operating-Systems/OS_FeatureUpdate_Counts.sql"),
        },
        "Memory Summary": {
            "source": "Hardware-Inventory/Memory_Summary.sql",
            "sql": read_sql("Hardware-Inventory/Memory_Summary.sql"),
        },
        "Device Models": {
            "source": "Hardware-Inventory/Device_Models.sql",
            "sql": read_sql("Hardware-Inventory/Device_Models.sql"),
        },
        "Client Version": {
            "source": "Client-Health/Client_Version.sql",
            "sql": read_sql("Client-Health/Client_Version.sql"),
        },
        "Update Compliance Summary": {
            "source": "Software-Updates/Update_Compliance_Summary.sql",
            "sql": read_sql("Software-Updates/Update_Compliance_Summary.sql"),
        },
        "Update Deployment Status": {
            "source": "Software-Updates/Update_Deployment_Status.sql",
            "sql": read_sql("Software-Updates/Update_Deployment_Status.sql"),
        },
        "Deployment Summary": {
            "source": "Applications/Deployment_Summary.sql",
            "sql": read_sql("Applications/Deployment_Summary.sql"),
        },
        "Application Deployment Status": {
            "source": "Applications/Application_Deployment_Status.sql",
            "sql": read_sql("Applications/Application_Deployment_Status.sql"),
        },
        "Server OS Versions": {
            "source": "Server/Server_OS_Versions.sql",
            "sql": read_sql("Server/Server_OS_Versions.sql"),
        },
    }

    # Add security aggregation queries
    for name, sql in SECURITY_QUERIES.items():
        sql_files[name] = {"source": "Dashboard (aggregation)", "sql": sql}

    # Build all sheets
    build_os_dashboard(wb)
    build_hardware_dashboard(wb)
    build_client_health(wb)
    build_update_compliance(wb)
    build_security_dashboard(wb)
    build_applications_dashboard(wb)
    build_server_dashboard(wb)
    build_queries_sheet(wb, sql_files)
    build_connection_sheet(wb)

    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"Workbook saved: {OUTPUT_FILE}")

    # Inject ODBC connections into xlsx zip
    conn_list = []
    for i, (name, info) in enumerate(sql_files.items(), start=1):
        conn_list.append({"id": i, "name": f"MECM - {name}", "sql": info["sql"]})

    inject_connections(OUTPUT_FILE, conn_list)
    print(f"Injected {len(conn_list)} data connections")
    print("Done.")


if __name__ == "__main__":
    main()
